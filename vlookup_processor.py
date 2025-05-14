from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import pandas as pd

def process_vlookup(file_a, file_b, sheet_name, start_line, last_row, skip_rows):
    wb = load_workbook(file_a)
    ws = wb[sheet_name]

    A_df = pd.read_excel(file_a, sheet_name=sheet_name, engine='openpyxl', header=6, skiprows=skip_rows)
    B_df = pd.read_excel(file_b, engine='openpyxl')

    B_last_row = B_df.index[-1] + 2
    lines = start_line

    while lines <= last_row:
        keyCol = 'A' + str(lines)
        outputCol = ws['G' + str(lines)]

        if isinstance(outputCol, MergedCell):
            lines += 1
            continue

        formula = f"=IFERROR(VLOOKUP({keyCol},'[O.xlsx]Sheet1'!$A$1:$B${B_last_row},2,0), 0)"
        outputCol.value = formula
        lines += 1

    output_path = "Processed_" + file_a.name
    wb.save(output_path)
    return output_path
